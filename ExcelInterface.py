import openpyxl
from shutil import copy2
from openpyxl.styles import PatternFill


import main


def initial():
    global wb, sheet, currentRow
    try:
        wb._archive.close()
    except:
        pass
    copy2("./init_excel/complete_schedule.xlsx", "./complete_schedule.xlsx")
    wb = openpyxl.load_workbook(filename='./complete_schedule.xlsx')
    sheet = wb['Schedule']
    currentRow = 2


def addRow(task):
    global currentRow
    sheet['A' + str(currentRow)] = task.name
    sheet['B' + str(currentRow)] = task.time
    sheet['C' + str(currentRow)] = task.deadline
    currentRow+=1
    wb.save("./complete_schedule.xlsx")


def makeRowCompleted(taskID):
    taskID+=2 #All starts with second line
    taskID = str(taskID)
    greenFill = PatternFill(start_color='FF008000', end_color='FF008000', fill_type='solid')
    sheet['A' + str(taskID)].fill = greenFill
    sheet['B' + str(taskID)].fill = greenFill
    sheet['C' + str(taskID)].fill = greenFill
    wb.save("./complete_schedule.xlsx")


def readFromList():
    initial()
    tasks_wb = openpyxl.load_workbook(filename='./tasks_list.xlsx')
    sheet1 = tasks_wb['Schedule']
    count = int(sheet1['A1'].value)
    tasks = []
    for i in range(2, count+2):
        tasks.append(main.Task(sheet1['A'+str(i)].value,
                               int(sheet1['B'+str(i)].value),
                               int(sheet1['C'+str(i)].value)))
    return tasks


wb = openpyxl.load_workbook(filename = './complete_schedule.xlsx')
sheet = wb['Schedule']
currentRow = 2
